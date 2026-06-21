"""
Report Service — PDF, Excel, CSV generation.
"""
import io
import csv
from typing import List
from datetime import datetime

from fastapi.responses import StreamingResponse


class ReportService:
    def generate_csv(self, data: List[dict], filename: str) -> StreamingResponse:
        if not data:
            output = io.StringIO()
            output.write("No data available")
            output.seek(0)
        else:
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
            output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    def generate_excel(self, data: List[dict], sheet_name: str, filename: str) -> StreamingResponse:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return self.generate_csv(data, filename.replace(".xlsx", ".csv"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            ws.append(["No data available"])
        else:
            headers = list(data[0].keys())
            ws.append(headers)

            # Header styling
            header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row in data:
                ws.append(list(row.values()))

            # Auto-width
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    def generate_pdf(self, title: str, data: List[dict], filename: str) -> StreamingResponse:
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            return self.generate_csv(data, filename.replace(".pdf", ".csv"))

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
        elements.append(Spacer(1, 20))

        if data:
            headers = list(data[0].keys())
            table_data = [headers] + [list(str(v) for v in row.values()) for row in data[:500]]

            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("PADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(table)

        doc.build(elements)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
